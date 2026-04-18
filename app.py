"""
智慧渔场管理系统增强版 - 带数据推送的Flask应用
Smart Fishery Management System v2.0 - Flask with Real-time Data
"""

import os
import sys
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
import random
import logging

# 硬件服务导入
try:
    from hardware_service import init_collector, start_collection, stop_collection, get_stats as get_hardware_stats
    from storage_service import init_storage_service, get_storage_service
    from custom_parser_template import CustomATKLoraParser
    
    # 初始化解析器
    parser = CustomATKLoraParser()
    
    def parse_hardware_data(raw_bytes):
        """解析硬件数据"""
        return parser.parse(raw_bytes)
    
    def get_parser():
        """获取解析器实例"""
        return parser
    
    HARDWARE_SUPPORT = True
except ImportError as e:
    print(f"[WARN] 硬件服务导入失败: {e}")
    HARDWARE_SUPPORT = False

# 配置日志
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============================================================
# Flask应用初始化
# ============================================================
app = Flask(__name__)

# 数据库配置
DB_HOST = '127.0.0.1'
DB_PORT = 3306
DB_USER = 'root'
DB_PASSWORD = '123456'
DB_NAME = 'smart_fishery_db'

DATABASE_URI = f'mysql+pymysql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}?charset=utf8mb4'

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ECHO'] = False
app.config['JSON_AS_ASCII'] = False

db = SQLAlchemy(app)

# ============================================================
# 登录安全配置
# ============================================================

# 登录尝试跟踪字典 {username: {'attempts': count, 'locked_until': timestamp}}
login_attempts = {}
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_DURATION = 600  # 秒（10分钟）


def is_account_locked(username):
    """检查账户是否被锁定"""
    if username not in login_attempts:
        return False
    
    attempt_data = login_attempts[username]
    if attempt_data['attempts'] < MAX_LOGIN_ATTEMPTS:
        return False
    
    # 检查锁定是否过期
    if datetime.utcnow() > attempt_data['locked_until']:
        login_attempts[username]['attempts'] = 0
        return False
    
    return True


def record_failed_attempt(username):
    """记录失败的登录尝试"""
    if username not in login_attempts:
        login_attempts[username] = {'attempts': 0, 'locked_until': None}
    
    login_attempts[username]['attempts'] += 1
    if login_attempts[username]['attempts'] >= MAX_LOGIN_ATTEMPTS:
        login_attempts[username]['locked_until'] = datetime.utcnow() + timedelta(seconds=LOCKOUT_DURATION)


def clear_login_attempts(username):
    """清除登录尝试计数"""
    if username in login_attempts:
        login_attempts[username]['attempts'] = 0
        login_attempts[username]['locked_until'] = None


# ============================================================
# 数据库模型
# ============================================================

class Pond(db.Model):
    __tablename__ = 'ponds'
    id = db.Column(db.Integer, primary_key=True)
    pond_name = db.Column(db.String(100), unique=True, nullable=False)
    fish_type = db.Column(db.String(50), nullable=False)
    fish_count = db.Column(db.Integer, default=0)
    volume = db.Column(db.Float, default=0.0)
    status = db.Column(db.String(20), default='正常')
    location = db.Column(db.String(255))
    default_supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id', ondelete='SET NULL'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'pond_name': self.pond_name,
            'fish_type': self.fish_type,
            'fish_count': self.fish_count,
            'volume': self.volume,
            'status': self.status,
            'location': self.location,
            'default_supplier_id': self.default_supplier_id,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class SensorData(db.Model):
    __tablename__ = 'sensor_data'
    id = db.Column(db.BigInteger, primary_key=True)
    pond_id = db.Column(db.Integer, db.ForeignKey('ponds.id'), nullable=False)
    temperature = db.Column(db.Float)
    ph_value = db.Column(db.Float)
    food_value = db.Column(db.Float)
    dissolved_oxygen = db.Column(db.Float)
    salinity = db.Column(db.Float)
    ammonia_nitrogen = db.Column(db.Float)
    nitrite_nitrogen = db.Column(db.Float)
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'pond_id': self.pond_id,
            'temperature': self.temperature,
            'ph_value': self.ph_value,
            'food_value': self.food_value,
            'dissolved_oxygen': self.dissolved_oxygen,
            'salinity': self.salinity,
            'ammonia_nitrogen': self.ammonia_nitrogen,
            'nitrite_nitrogen': self.nitrite_nitrogen,
            'recorded_at': self.recorded_at.isoformat() if self.recorded_at else None
        }


class Device(db.Model):
    __tablename__ = 'devices'
    id = db.Column(db.Integer, primary_key=True)
    pond_id = db.Column(db.Integer, db.ForeignKey('ponds.id'), nullable=False)
    device_name = db.Column(db.String(100), nullable=False)
    device_type = db.Column(db.String(50), nullable=False)
    device_model = db.Column(db.String(100))
    status = db.Column(db.String(20), default='离线')
    power_consumption = db.Column(db.Float, default=0.0)
    last_active = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'pond_id': self.pond_id,
            'device_name': self.device_name,
            'device_type': self.device_type,
            'device_model': self.device_model,
            'status': self.status,
            'power_consumption': self.power_consumption,
            'last_active': self.last_active.isoformat() if self.last_active else None
        }


class DeviceLog(db.Model):
    __tablename__ = 'device_logs'
    id = db.Column(db.BigInteger, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('devices.id'), nullable=False)
    pond_id = db.Column(db.Integer, db.ForeignKey('ponds.id'), nullable=False)
    action = db.Column(db.String(50), nullable=False)
    operator = db.Column(db.String(100))
    previous_state = db.Column(db.String(20))
    current_state = db.Column(db.String(20))
    details = db.Column(db.Text)
    log_time = db.Column(db.DateTime, default=datetime.utcnow)


# ============================================================
# 用户认证模型（任务4）
# ============================================================
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), default='user')
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id', ondelete='SET NULL'), nullable=True)
    email = db.Column(db.String(100), unique=True, nullable=True)
    full_name = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    last_login = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'role': self.role,
            'supplier_id': self.supplier_id,
            'email': self.email,
            'full_name': self.full_name,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


# ============================================================
# 供应商管理模型（任务5）
# ============================================================

class Supplier(db.Model):
    __tablename__ = 'suppliers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    contact_person = db.Column(db.String(50))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.String(255))
    registration_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='active')
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # 关系
    products = db.relationship('SeedlingProduct', backref='supplier', lazy='dynamic', cascade='all, delete-orphan')
    users = db.relationship('User', backref='supplier_obj', lazy='dynamic')
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'contact_person': self.contact_person,
            'phone': self.phone,
            'email': self.email,
            'address': self.address,
            'registration_date': self.registration_date.isoformat() if self.registration_date else None,
            'status': self.status,
            'notes': self.notes,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class SeedlingProduct(db.Model):
    __tablename__ = 'seedling_products'
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id', ondelete='CASCADE'), nullable=False)
    product_name = db.Column(db.String(100), nullable=False)
    species = db.Column(db.String(50), nullable=False)
    grade = db.Column(db.String(50))
    unit_price = db.Column(db.Float, default=0.0)
    cost_price = db.Column(db.Float)
    growth_cycle_days = db.Column(db.Integer)
    survival_rate = db.Column(db.Float)
    image_url = db.Column(db.String(255))
    description = db.Column(db.Text)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'supplier_id': self.supplier_id,
            'product_name': self.product_name,
            'species': self.species,
            'grade': self.grade,
            'unit_price': self.unit_price,
            'cost_price': self.cost_price,
            'growth_cycle_days': self.growth_cycle_days,
            'survival_rate': self.survival_rate,
            'image_url': self.image_url,
            'description': self.description,
            'is_active': self.is_active,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class SeedlingInventory(db.Model):
    __tablename__ = 'seedling_inventory'
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id', ondelete='CASCADE'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('seedling_products.id', ondelete='CASCADE'), nullable=False)
    quantity = db.Column(db.Integer, default=0)
    last_updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = db.Column(db.String(50))
    
    def to_dict(self):
        return {
            'id': self.id,
            'supplier_id': self.supplier_id,
            'product_id': self.product_id,
            'quantity': self.quantity,
            'last_updated_at': self.last_updated_at.isoformat() if self.last_updated_at else None,
            'updated_by': self.updated_by
        }


class PurchaseOrder(db.Model):
    __tablename__ = 'purchase_orders'
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id', ondelete='RESTRICT'), nullable=False)
    pond_id = db.Column(db.Integer, db.ForeignKey('ponds.id', ondelete='RESTRICT'), nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id', ondelete='RESTRICT'), nullable=False)
    order_date = db.Column(db.DateTime, default=datetime.utcnow)
    expected_delivery_date = db.Column(db.Date)
    actual_delivery_date = db.Column(db.Date)
    status = db.Column(db.String(50), default='draft')
    total_amount = db.Column(db.Float, default=0.0)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # 关系
    items = db.relationship('OrderItem', backref='order', lazy='dynamic', cascade='all, delete-orphan')
    creator = db.relationship('User', backref='created_orders')
    pond_ref = db.relationship('Pond', backref='purchase_orders')
    supplier_ref = db.relationship('Supplier', backref='purchase_orders')
    
    def to_dict(self):
        return {
            'id': self.id,
            'supplier_id': self.supplier_id,
            'pond_id': self.pond_id,
            'created_by': self.created_by,
            'order_date': self.order_date.isoformat() if self.order_date else None,
            'expected_delivery_date': self.expected_delivery_date.isoformat() if self.expected_delivery_date else None,
            'actual_delivery_date': self.actual_delivery_date.isoformat() if self.actual_delivery_date else None,
            'status': self.status,
            'total_amount': self.total_amount,
            'notes': self.notes,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class OrderItem(db.Model):
    __tablename__ = 'order_items'
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('purchase_orders.id', ondelete='CASCADE'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('seedling_products.id', ondelete='RESTRICT'), nullable=False)
    quantity = db.Column(db.Integer, default=0)
    unit_price = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系
    product = db.relationship('SeedlingProduct', backref='order_items')
    
    def to_dict(self):
        return {
            'id': self.id,
            'order_id': self.order_id,
            'product_id': self.product_id,
            'quantity': self.quantity,
            'unit_price': self.unit_price,
            'subtotal': self.quantity * self.unit_price,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class WaterQualityThreshold(db.Model):
    """水质告警阈值配置 (Task3)"""
    __tablename__ = 'water_quality_thresholds'
    id = db.Column(db.Integer, primary_key=True)
    parameter_name = db.Column(db.String(50), nullable=False, unique=True)  # 温度、pH、溶氧等
    parameter_key = db.Column(db.String(50), nullable=False)  # temperature, ph_value, dissolved_oxygen等
    min_value = db.Column(db.Float)
    max_value = db.Column(db.Float)
    warning_level = db.Column(db.String(20), default='warning')  # warning, critical
    unit = db.Column(db.String(20))
    description = db.Column(db.String(255))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'id': self.id,
            'parameter_name': self.parameter_name,
            'parameter_key': self.parameter_key,
            'min_value': self.min_value,
            'max_value': self.max_value,
            'warning_level': self.warning_level,
            'unit': self.unit,
            'description': self.description,
            'is_active': self.is_active,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


class SystemLog(db.Model):
    """系统操作审计日志 (Task4)"""
    __tablename__ = 'system_logs'
    id = db.Column(db.BigInteger, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id', ondelete='SET NULL'), nullable=True)
    username = db.Column(db.String(100))
    action = db.Column(db.String(50), nullable=False)  # create, update, delete, control, login
    resource_type = db.Column(db.String(50), nullable=False)  # Pond, Device, SeedlingProduct等
    resource_id = db.Column(db.Integer)
    resource_name = db.Column(db.String(255))
    old_value = db.Column(db.Text)
    new_value = db.Column(db.Text)
    ip_address = db.Column(db.String(50))
    details = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    
    def to_dict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'username': self.username,
            'action': self.action,
            'resource_type': self.resource_type,
            'resource_id': self.resource_id,
            'resource_name': self.resource_name,
            'old_value': self.old_value,
            'new_value': self.new_value,
            'ip_address': self.ip_address,
            'details': self.details,
            'created_at': self.created_at.isoformat() if self.created_at else None
        }


# ============================================================
# 硬件数据采集初始化
# ============================================================

def init_hardware_collection():
    """初始化硬件数据采集"""
    if not HARDWARE_SUPPORT:
        print("[WARN] 硬件支持未启用")
        return
    
    try:
        # 初始化存储服务 (传递Flask应用对象)
        storage = init_storage_service(app, db, Pond, SensorData)
        
        # 硬件数据始终更新一号池（ID=1）
        pond_id = 1
        
        # 检查一号池是否存在
        with app.app_context():
            pond = Pond.query.get(1)
            if pond:
                logger.info(f"[OK] 硬件采集已绑定到: {pond.pond_name} (ID=1)")
            else:
                logger.warning("[WARN] 一号池（ID=1）不存在！")
        
        storage.set_default_pond(pond_id)
        
        # 数据回调函数 - 接收硬件数据后的处理
        def on_hardware_data(raw_bytes, timestamp):
            """硬件数据回调 - 解析并存储数据"""
            try:
                # 尝试解析硬件数据
                if parse_hardware_data is None:
                    logger.warning("解析器不可用")
                    return
                
                parsed_data = parse_hardware_data(raw_bytes)
                
                storage = get_storage_service()
                if storage and parsed_data:
                    # 成功解析数据，保存到数据库
                    storage.store_parsed_data(
                        pond_id=pond_id,
                        temperature=parsed_data.get('temperature'),
                        ph_value=parsed_data.get('ph_value'),
                        food_value=parsed_data.get('food_value'),
                        dissolved_oxygen=parsed_data.get('dissolved_oxygen'),
                        salinity=parsed_data.get('salinity'),
                        ammonia_nitrogen=parsed_data.get('ammonia_nitrogen'),
                        nitrite_nitrogen=parsed_data.get('nitrite_nitrogen'),
                        timestamp=timestamp
                    )
                    logger.debug(f"✓ 传感器数据已保存: T={parsed_data.get('temperature'):.2f}°C, Food={parsed_data.get('food_value')}")

                
            except Exception as e:
                logger.error(f"数据回调处理失败: {e}")
        
        # 初始化采集器
        collector = init_collector(port='COM12', baudrate=9600, data_callback=on_hardware_data)
        
        # 启动采集
        collector.start()
        print("[OK] 硬件数据采集已启动")
        
        # 输出解析器状态
        if HARDWARE_SUPPORT:
            parser = get_parser()
            print(f"📊 解析器就绪: {parser.__class__.__name__}")
        
    except Exception as e:
        print(f"[ERR] 硬件初始化失败: {e}")


# 应用启动事件
@app.before_request
def before_first_request():
    """首次请求前的初始化"""
    pass


def shutdown_hardware():
    """应用关闭时停止采集"""
    stop_collection()
    print("⏹️  硬件采集已停止")


app.teardown_appcontext(lambda exc: None)  # Placeholder

# ============================================================
# 认证和权限检查（任务4）
# ============================================================
from functools import wraps

# 设置 session 密钥
app.secret_key = 'smartfishery-secret-key-2026'

def login_required(f):
    """权限检查装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def api_login_required(f):
    """API 权限检查装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'status': 'error', 'message': '未认证，请登录'}), 401
        return f(*args, **kwargs)
    return decorated_function


# ============================================================
# 权限装饰器（任务6）
# ============================================================

def role_required(allowed_roles):
    """角色权限检查装饰器"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'status': 'error', 'message': '未认证，请登录'}), 401
            
            user = User.query.get(session['user_id'])
            if not user or user.role not in allowed_roles:
                return jsonify({'status': 'error', 'message': '您没有权限访问此资源'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def supplier_scope_check(f):
    """供应商数据隔离检查装饰器 - 确保供应商只能访问自己的数据"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'status': 'error', 'message': '未认证，请登录'}), 401
        
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'status': 'error', 'message': '用户不存在'}), 401
        
        # 管理员可以访问所有数据
        if user.role == 'admin':
            return f(*args, **kwargs)
        
        # 供应商只能访问自己的数据
        if user.role == 'supplier' and user.supplier_id:
            # 将supplier_id注入到kwargs中供路由函数使用
            kwargs['_supplier_id'] = user.supplier_id
            return f(*args, **kwargs)
        
        return jsonify({'status': 'error', 'message': '没有权限访问此数据'}), 403
    return decorated_function


# 登录页面
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        identity = request.form.get('identity', 'admin')  # 获取选择的身份
        
        # 检查账户是否被锁定
        if is_account_locked(username):
            remaining_time = int((login_attempts[username]['locked_until'] - datetime.utcnow()).total_seconds())
            error_msg = f'账户已被锁定，请在{remaining_time}秒后再试'
            return render_template('login.html', error=error_msg)
        
        user = User.query.filter_by(username=username).first()
        # 简单密码验证（生产环境应使用哈希）
        if user and user.password_hash == password:
            # 验证选择的身份是否与用户角色匹配
            if identity == 'admin' and user.role != 'admin':
                record_failed_attempt(username)
                return render_template('login.html', error='该账号不是管理员账号，请选择"鱼苗供应商"身份')
            elif identity == 'supplier' and user.role != 'supplier':
                record_failed_attempt(username)
                return render_template('login.html', error='该账号不是供应商账号，请选择"渔场管理员"身份')
            
            # 清除登录尝试计数
            clear_login_attempts(username)
            
            # 更新最后登录时间
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            session['supplier_id'] = user.supplier_id
            
            # 记录登录日志
            record_audit_log('login', 'User', user.id, details=f'用户{username}登录')
            
            # 根据角色重定向到不同的仪表板
            if user.role == 'admin':
                return redirect(url_for('index'))
            elif user.role == 'supplier':
                return redirect(url_for('supplier_dashboard'))
            else:
                return redirect(url_for('index'))
        else:
            record_failed_attempt(username)
            attempts_remaining = MAX_LOGIN_ATTEMPTS - login_attempts.get(username, {}).get('attempts', 0)
            if attempts_remaining > 0:
                error_msg = f'用户名或密码错误 (还有{attempts_remaining}次尝试机会)'
            else:
                error_msg = '登录尝试过多，账户已被锁定'
            return render_template('login.html', error=error_msg)
    
    return render_template('login.html')

# 登出
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ============================================================
# 页面路由
# ============================================================

@app.route('/')
@login_required
def index():
    try:
        pond_count = Pond.query.count()
        total_fish_count = db.session.query(db.func.sum(Pond.fish_count)).scalar() or 0
        total_devices = Device.query.count()
        online_devices = Device.query.filter_by(status='在线').count()
        running_devices = Device.query.filter_by(status='运行中').count()
        
        devices_list = Device.query.all()
        devices = [{'device_name': d.device_name, 'device_type': d.device_type, 'status': d.status} for d in devices_list]
        
        # 从数据库获取第一个鱼池的最新水质数据
        first_pond = Pond.query.first()
        if first_pond:
            latest_sensor = SensorData.query.filter_by(pond_id=first_pond.id).order_by(SensorData.recorded_at.desc()).first()
        else:
            latest_sensor = None
            
        if latest_sensor:
            water_quality_data = {
                'values': [
                    latest_sensor.temperature or 0,
                    latest_sensor.ph_value or 0,
                    latest_sensor.food_value or 0,
                    latest_sensor.dissolved_oxygen or 0,
                    latest_sensor.salinity or 0
                ]
            }
        else:
            water_quality_data = {'values': [0, 0, 0, 0, 0]}
        
        # 设备状态统计
        device_status_data = {
            'online': online_devices,
            'offline': max(0, total_devices - online_devices),
            'running': running_devices
        }
        
        # 从数据库获取第一个鱼池的最近12条传感器数据（如果有的话）
        recent_sensors = []
        first_pond = Pond.query.first()
        if first_pond:
            recent_sensors = SensorData.query.filter_by(pond_id=first_pond.id).order_by(SensorData.recorded_at.desc()).limit(12).all()
        
        if recent_sensors:
            recent_sensors.reverse()
            recent_data = {
                'temperature': [s.temperature or 0 for s in recent_sensors],
                'oxygen': [s.dissolved_oxygen or 0 for s in recent_sensors]
            }
        else:
            # 如果数据库中没有数据，使用演示数据
            recent_data = {
                'temperature': [24+i*0.5 for i in range(12)],
                'oxygen': [8.5-i*0.1 for i in range(12)]
            }
        
        return render_template('dashboard.html',
                             pond_count=pond_count,
                             total_fish_count=int(total_fish_count),
                             total_devices=total_devices,
                             online_devices=online_devices,
                             devices=devices[:8],
                             water_quality_data=water_quality_data,
                             device_status_data=device_status_data,
                             recent_data=recent_data,
                             now=datetime.now())
    except Exception as e:
        print(f'Dashboard error: {e}')
        return render_template('dashboard.html',
                             pond_count=0, total_fish_count=0, total_devices=0,
                             online_devices=0, devices=[], water_quality_data={'values': [0,0,0,0,0]},
                             device_status_data={'online': 0, 'offline': 0, 'running': 0},
                             recent_data={'temperature': [0]*12, 'oxygen': [0]*12},
                             now=datetime.now())


@app.route('/ponds')
@login_required
def ponds_page():
    try:
        ponds = Pond.query.all()
        return render_template('ponds.html', ponds=ponds)
    except Exception as e:
        print(f'Ponds error: {e}')
        return render_template('ponds.html', ponds=[])


@app.route('/water-quality')
@login_required
def water_quality_page():
    try:
        ponds = Pond.query.all()
        
        # 获取各鱼池的最新水质数据
        pond_quality_data = {}
        for pond in ponds:
            latest = SensorData.query.filter_by(pond_id=pond.id).order_by(SensorData.recorded_at.desc()).first()
            if latest:
                pond_quality_data[pond.id] = {
                    'temperature': latest.temperature or 0,
                    'ph_value': latest.ph_value or 0,
                    'food_value': latest.food_value or 0,
                    'dissolved_oxygen': latest.dissolved_oxygen or 0,
                    'salinity': latest.salinity or 0,
                    'ammonia_nitrogen': latest.ammonia_nitrogen or 0,
                    'nitrite_nitrogen': latest.nitrite_nitrogen or 0,
                    'recorded_at': latest.recorded_at.strftime('%Y-%m-%d %H:%M:%S') if latest.recorded_at else '未知'
                }
        
        return render_template('water_quality.html', 
                             ponds=ponds,
                             pond_quality_data=pond_quality_data)
    except Exception as e:
        print(f'Water quality error: {e}')
        return render_template('water_quality.html', ponds=[], pond_quality_data={})


@app.route('/devices')
@login_required
def devices_page():
    try:
        devices = Device.query.all()
        return render_template('devices.html', devices=devices)
    except Exception as e:
        print(f'Devices error: {e}')
        return render_template('devices.html', devices=[])


# ============================================================
# API路由
# ============================================================

@app.route('/api/dashboard-refresh', methods=['GET'])
def dashboard_refresh():
    """刷新仪表板所有数据"""
    try:
        pond_count = Pond.query.count()
        total_fish_count = db.session.query(db.func.sum(Pond.fish_count)).scalar() or 0
        total_devices = Device.query.count()
        online_devices = Device.query.filter_by(status='在线').count()
        running_devices = Device.query.filter_by(status='运行中').count()
        
        # 获取第一个鱼池的最新水质数据
        first_pond = Pond.query.first()
        water_quality = None
        if first_pond:
            latest_sensor = SensorData.query.filter_by(pond_id=first_pond.id).order_by(SensorData.recorded_at.desc()).first()
            if latest_sensor:
                water_quality = {
                    'temperature': latest_sensor.temperature or 0,
                    'ph_value': latest_sensor.ph_value or 0,
                    'dissolved_oxygen': latest_sensor.dissolved_oxygen or 0,
                    'salinity': latest_sensor.salinity or 0
                }
        
        # 获取最新的设备列表
        devices_list = Device.query.limit(8).all()
        devices = [{'device_name': d.device_name, 'device_type': d.device_type, 'status': d.status} for d in devices_list]
        
        # 获取第一个鱼池的最近12条传感器数据
        recent_sensors = []
        if first_pond:
            recent_sensors = SensorData.query.filter_by(pond_id=first_pond.id).order_by(SensorData.recorded_at.desc()).limit(12).all()
        
        recent_data = {
            'temperature': [s.temperature or 0 for s in reversed(recent_sensors)],
            'oxygen': [s.dissolved_oxygen or 0 for s in reversed(recent_sensors)]
        }
        
        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'data': {
                'pond_count': pond_count,
                'total_fish_count': int(total_fish_count),
                'total_devices': total_devices,
                'online_devices': online_devices,
                'running_devices': running_devices,
                'water_quality': water_quality,
                'devices': devices,
                'recent_data': recent_data
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/water-quality/all-ponds', methods=['GET'])
@api_login_required
def water_quality_all_ponds():
    """获取所有鱼池的水质数据汇总"""
    try:
        ponds = Pond.query.all()
        
        # 获取每个鱼池的最新水质数据
        pond_data_list = []
        temp_values = []
        ph_values = []
        oxygen_values = []
        abnormal_count = 0
        
        for pond in ponds:
            latest = SensorData.query.filter_by(pond_id=pond.id).order_by(SensorData.recorded_at.desc()).first()
            if latest:
                temp = latest.temperature or 0
                ph = latest.ph_value or 0
                oxygen = latest.dissolved_oxygen or 0
                salinity = latest.salinity or 0
                
                # 收集数据用于计算平均值
                if temp > 0:
                    temp_values.append(temp)
                if ph > 0:
                    ph_values.append(ph)
                if oxygen > 0:
                    oxygen_values.append(oxygen)
                
                # 判断状态
                status = 'normal'
                if temp < 20 or temp > 35:
                    status = 'warning'
                if ph < 6.0 or ph > 8.5:
                    status = 'warning'
                if oxygen < 5:
                    status = 'danger'
                    abnormal_count += 1
                
                pond_data_list.append({
                    'pond_id': pond.id,
                    'pond_name': pond.pond_name,
                    'fish_type': pond.fish_type or '未知',
                    'fish_count': pond.fish_count or 0,
                    'temperature': round(temp, 2),
                    'ph_value': round(ph, 2),
                    'dissolved_oxygen': round(oxygen, 2),
                    'salinity': round(salinity, 2),
                    'food_value': round(latest.food_value or 0, 2),
                    'ammonia_nitrogen': round(latest.ammonia_nitrogen or 0, 2),
                    'nitrite_nitrogen': round(latest.nitrite_nitrogen or 0, 2),
                    'recorded_at': latest.recorded_at.strftime('%Y-%m-%d %H:%M:%S') if latest.recorded_at else '未知',
                    'status': status
                })
        
        # 计算平均值
        avg_temperature = round(sum(temp_values) / len(temp_values), 2) if temp_values else 0
        avg_ph_value = round(sum(ph_values) / len(ph_values), 2) if ph_values else 0
        avg_dissolved_oxygen = round(sum(oxygen_values) / len(oxygen_values), 2) if oxygen_values else 0
        
        return jsonify({
            'status': 'success',
            'summary': {
                'total_ponds': len(ponds),
                'avg_temperature': avg_temperature,
                'avg_ph_value': avg_ph_value,
                'avg_dissolved_oxygen': avg_dissolved_oxygen,
                'abnormal_count': abnormal_count
            },
            'ponds': pond_data_list
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/water-quality-refresh/<int:pond_id>', methods=['GET'])
def water_quality_refresh(pond_id):
    """刷新指定鱼池的水质数据"""
    try:
        data = SensorData.query.filter_by(pond_id=pond_id).order_by(SensorData.recorded_at.desc()).first()
        if not data:
            return jsonify({'status': 'error', 'message': '该鱼池暂无数据'}), 404
        
        return jsonify({
            'status': 'success',
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'data': {
                'temperature': data.temperature or 0,
                'ph_value': data.ph_value or 0,
                'food_value': data.food_value or 0,
                'dissolved_oxygen': data.dissolved_oxygen or 0,
                'salinity': data.salinity or 0,
                'ammonia_nitrogen': data.ammonia_nitrogen or 0,
                'nitrite_nitrogen': data.nitrite_nitrogen or 0,
                'recorded_at': data.recorded_at.strftime('%H:%M:%S') if data.recorded_at else '未知'
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/ponds/add', methods=['POST'])
def add_pond():
    """添加新鱼池"""
    try:
        data = request.get_json()
        pond_name = data.get('pond_name')
        fish_type = data.get('fish_type')
        fish_count = data.get('fish_count', 0)
        volume = data.get('volume', 0)
        location = data.get('location', '')
        
        # 验证必填字段
        if not pond_name or not fish_type:
            return jsonify({'status': 'error', 'message': '鱼池名称和鱼类类型为必填项'}), 400
        
        # 检查鱼池名称是否已存在
        existing_pond = Pond.query.filter_by(pond_name=pond_name).first()
        if existing_pond:
            return jsonify({'status': 'error', 'message': '该鱼池名称已存在'}), 400
        
        pond = Pond(
            pond_name=pond_name,
            fish_type=fish_type,
            fish_count=int(fish_count),
            volume=float(volume),
            location=location,
            status='正常'
        )
        db.session.add(pond)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': '鱼池添加成功',
            'data': pond.to_dict()
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/ponds/<int:pond_id>/edit', methods=['POST'])
def edit_pond(pond_id):
    """编辑鱼池信息"""
    try:
        pond = Pond.query.get(pond_id)
        if not pond:
            return jsonify({'status': 'error', 'message': '鱼池不存在'}), 404
        
        data = request.get_json()
        pond.pond_name = data.get('pond_name', pond.pond_name)
        pond.fish_type = data.get('fish_type', pond.fish_type)
        pond.fish_count = int(data.get('fish_count', pond.fish_count))
        pond.volume = float(data.get('volume', pond.volume))
        pond.location = data.get('location', pond.location)
        
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': '鱼池信息更新成功',
            'data': pond.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/ponds/<int:pond_id>/delete', methods=['DELETE'])
def delete_pond(pond_id):
    """删除鱼池"""
    try:
        pond = Pond.query.get(pond_id)
        if not pond:
            return jsonify({'status': 'error', 'message': '鱼池不存在'}), 404
        
        # 删除相关数据
        SensorData.query.filter_by(pond_id=pond_id).delete()
        Device.query.filter_by(pond_id=pond_id).delete()
        DeviceLog.query.filter_by(pond_id=pond_id).delete()
        db.session.delete(pond)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': '鱼池删除成功'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({'status': 'success', 'message': '系统正常运行'}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/ponds', methods=['GET'])
def get_ponds():
    try:
        ponds = Pond.query.all()
        return jsonify({'status': 'success', 'data': [p.to_dict() for p in ponds]}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/ponds/<int:pond_id>', methods=['GET'])
def get_pond(pond_id):
    try:
        pond = Pond.query.get(pond_id)
        if not pond:
            return jsonify({'status': 'error', 'message': '鱼池不存在'}), 404
        return jsonify({'status': 'success', 'data': pond.to_dict()}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sensor-data/<int:pond_id>', methods=['GET'])
def get_sensor_data(pond_id):
    """
    获取指定鱼池的传感器数据（支持分页和时间范围过滤）
    查询参数：page（页码），per_page（每页数量），hours（过去N小时内的数据）
    """
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        hours = request.args.get('hours', None, type=int)
        
        query = SensorData.query.filter_by(pond_id=pond_id)
        
        # 支持时间范围过滤
        if hours:
            from datetime import timedelta
            cutoff_time = datetime.utcnow() - timedelta(hours=hours)
            query = query.filter(SensorData.recorded_at >= cutoff_time)
        
        # 使用分页并按时间倒序排列
        pagination = query.order_by(SensorData.recorded_at.desc()).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
        
        return jsonify({
            'status': 'success',
            'data': [d.to_dict() for d in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/devices/<int:pond_id>', methods=['GET'])
def get_devices(pond_id):
    """
    获取指定鱼池的设备列表（支持分页）
    查询参数：page（页码，默认1），per_page（每页数量，默认20）
    """
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # 使用分页查询，优化大数据集性能
        pagination = Device.query.filter_by(pond_id=pond_id).paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        return jsonify({
            'status': 'success', 
            'data': [d.to_dict() for d in pagination.items],
            'total': pagination.total,
            'pages': pagination.pages,
            'current_page': page
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/devices/add', methods=['POST'])
@api_login_required
def add_device():
    """添加新设备 (Task2)"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可添加设备'}), 403
        
        data = request.get_json()
        
        # 验证必要字段
        if not data.get('device_name'):
            return jsonify({'status': 'error', 'message': '设备名称不能为空'}), 400
        if not data.get('device_type'):
            return jsonify({'status': 'error', 'message': '设备类型不能为空'}), 400
        if not data.get('pond_id'):
            return jsonify({'status': 'error', 'message': '所属鱼塘不能为空'}), 400
        
        # 检查鱼塘是否存在
        pond = Pond.query.get(data['pond_id'])
        if not pond:
            return jsonify({'status': 'error', 'message': '选择的鱼塘不存在'}), 404
        
        # 创建新设备
        new_device = Device(
            pond_id=data['pond_id'],
            device_name=data['device_name'],
            device_type=data['device_type'],
            device_model=data.get('device_model', '标准型'),
            status='离线',  # 新设备默认离线
            power_consumption=float(data.get('power_consumption', 0)),
            last_active=datetime.utcnow(),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        db.session.add(new_device)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'设备"{data["device_name"]}"添加成功',
            'device_id': new_device.id
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/devices/<int:device_id>/control', methods=['POST'])
def control_device(device_id):
    try:
        action = request.json.get('action')
        device = Device.query.get(device_id)
        if not device:
            return jsonify({'status': 'error', 'message': '设备不存在'}), 404
        
        device.status = '运行中' if action == 'start' else '停止'
        device.last_active = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': f'设备{action}成功'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/dashboard-stats', methods=['GET'])
def get_dashboard_stats():
    """
    获取仪表板统计数据（优化：使用数据库级聚合）
    """
    try:
        # 一次查询获取多个统计值，避免多次查询
        from sqlalchemy import func
        
        stats_query = db.session.query(
            func.count(Pond.id).label('pond_count'),
            func.count(Device.id).label('total_devices'),
            func.sum(Pond.fish_count).label('total_fish_count'),
            func.sum(db.case(
                (Device.status == '在线', 1),
                else_=0
            )).label('online_devices'),
            func.sum(db.case(
                (Device.status == '运行中', 1),
                else_=0
            )).label('running_devices')
        ).outerjoin(Device).first()
        
        stats = {
            'pond_count': stats_query.pond_count or 0,
            'total_devices': stats_query.total_devices or 0,
            'total_fish_count': int(stats_query.total_fish_count or 0),
            'online_devices': stats_query.online_devices or 0,
            'running_devices': stats_query.running_devices or 0
        }
        
        return jsonify({'status': 'success', 'data': stats}), 200
    except Exception as e:
        # 如果聚合查询失败，降级到多个查询
        try:
            stats = {
                'pond_count': Pond.query.count(),
                'total_fish_count': db.session.query(db.func.sum(Pond.fish_count)).scalar() or 0,
                'total_devices': Device.query.count(),
                'online_devices': Device.query.filter_by(status='在线').count(),
                'running_devices': Device.query.filter_by(status='运行中').count()
            }
            return jsonify({'status': 'success', 'data': stats}), 200
        except Exception as e2:
            return jsonify({'status': 'error', 'message': str(e2)}), 500


@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """获取系统统计数据（优化：复用仪表板统计接口）"""
    try:
        # 复用仪表板统计，避免重复查询
        stats_response = get_dashboard_stats()
        if stats_response[1] == 200:
            return stats_response
        
        # 如果失败，降级方案
        stats = {
            'pond_count': Pond.query.count(),
            'total_fish_count': int(db.session.query(db.func.sum(Pond.fish_count)).scalar() or 0),
            'total_devices': Device.query.count(),
            'online_devices': Device.query.filter_by(status='在线').count()
        }
        return jsonify({'status': 'success', 'data': stats}), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 硬件监控API
# ============================================================

@app.route('/api/hardware/status', methods=['GET'])
def get_hardware_status():
    """获取硬件采集状态"""
    try:
        if not HARDWARE_SUPPORT:
            return jsonify({'status': 'error', 'message': '硬件支持未启用'}), 500
        
        stats = get_hardware_stats()
        
        # 添加解析器统计
        if 'data' in stats and HARDWARE_SUPPORT:
            try:
                parser = get_parser()
                stats['data']['parser_stats'] = parser.get_stats()
            except:
                pass
        
        return jsonify({
            'status': 'success',
            'data': stats
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/hardware/start', methods=['POST'])
def hardware_start():
    """启动硬件采集"""
    try:
        if not HARDWARE_SUPPORT:
            return jsonify({'status': 'error', 'message': '硬件支持未启用'}), 500
        
        start_collection()
        return jsonify({
            'status': 'success',
            'message': '硬件采集已启动'
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/hardware/stop', methods=['POST'])
def hardware_stop():
    """停止硬件采集"""
    try:
        stop_collection()
        return jsonify({
            'status': 'success',
            'message': '硬件采集已停止'
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sensor-data/latest/<int:pond_id>', methods=['GET'])
def get_latest_sensor_data(pond_id):
    """获取指定鱼池的最新传感器数据（使用索引优化查询）"""
    try:
        # 使用组合索引 idx_pond_recorded (pond_id, recorded_at) 优化查询
        # order_by + first() 比 first() 后再 order_by 更高效
        data = SensorData.query.filter_by(pond_id=pond_id)\
            .order_by(SensorData.recorded_at.desc())\
            .first()
        
        if not data:
            return jsonify({'status': 'success', 'data': None}), 200
        
        return jsonify({
            'status': 'success',
            'data': data.to_dict()
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 前端页面路由（任务7-9）
# ============================================================

@app.route('/supplier-dashboard', methods=['GET'])
@login_required
def supplier_dashboard():
    """供应商仪表板"""
    try:
        if session.get('role') != 'supplier':
            return redirect(url_for('login_page'))
        
        supplier_id = session.get('supplier_id')
        supplier = Supplier.query.get(supplier_id)
        
        return render_template('supplier-dashboard.html', 
                               username=session.get('username'),
                               role=session.get('role'),
                               supplier_name=supplier.name if supplier else '供应商')
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/supplier-products', methods=['GET'])
@login_required
def supplier_products():
    """供应商产品管理"""
    try:
        if session.get('role') != 'supplier':
            return redirect(url_for('login_page'))
        
        return render_template('supplier-products.html', 
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/supplier-orders', methods=['GET'])
@login_required
def supplier_orders():
    """供应商订单管理"""
    try:
        if session.get('role') != 'supplier':
            return redirect(url_for('login_page'))
        
        return render_template('supplier-orders.html', 
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/supplier-stats', methods=['GET'])
@login_required
def supplier_stats():
    """供应商财务统计"""
    try:
        if session.get('role') != 'supplier':
            return redirect(url_for('login_page'))
        
        return render_template('supplier-stats.html', 
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/seedling-management', methods=['GET'])
@login_required
def seedling_management_page():
    """管理员鱼苗管理中心"""
    try:
        if session.get('role') != 'admin':
            return redirect(url_for('login_page'))
        
        return render_template('seedling-management.html', 
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


# ============================================================
# 水质告警阈值配置 API (Task3)
# ============================================================

@app.route('/api/thresholds', methods=['GET'])
@api_login_required
def get_thresholds():
    """获取所有水质告警阈值"""
    try:
        thresholds = WaterQualityThreshold.query.filter_by(is_active=True).all()
        return jsonify({
            'status': 'success',
            'data': [t.to_dict() for t in thresholds]
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/thresholds/update', methods=['POST'])
@api_login_required
def update_thresholds():
    """更新水质告警阈值"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可以修改阈值'}), 403
        
        data = request.get_json()
        threshold_id = data.get('id')
        
        threshold = WaterQualityThreshold.query.get(threshold_id)
        if not threshold:
            return jsonify({'status': 'error', 'message': '阈值不存在'}), 404
        
        # 更新字段
        if 'min_value' in data:
            threshold.min_value = float(data['min_value']) if data['min_value'] is not None else None
        if 'max_value' in data:
            threshold.max_value = float(data['max_value']) if data['max_value'] is not None else None
        if 'warning_level' in data:
            threshold.warning_level = data['warning_level']
        if 'is_active' in data:
            threshold.is_active = data['is_active']
        
        threshold.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': '阈值更新成功',
            'data': threshold.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/water-quality-thresholds', methods=['GET'])
@login_required
def water_quality_thresholds():
    """水质告警阈值配置页面"""
    try:
        if session.get('role') != 'admin':
            return redirect(url_for('login_page'))
        
        return render_template('water-quality-thresholds.html',
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


# ============================================================
# 系统操作日志与审计 API (Task4)
# ============================================================

def record_audit_log(action, resource_type, resource_id=None, resource_name=None, 
                     old_value=None, new_value=None, details=None):
    """记录系统操作审计日志"""
    try:
        user_id = session.get('user_id')
        username = session.get('username')
        ip_address = request.remote_addr
        
        log = SystemLog(
            user_id=user_id,
            username=username,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            resource_name=resource_name,
            old_value=old_value,
            new_value=new_value,
            ip_address=ip_address,
            details=details,
            created_at=datetime.utcnow()
        )
        
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"[WARN] 审计日志记录失败: {e}")
        db.session.rollback()


@app.route('/system-logs', methods=['GET'])
@login_required
def system_logs_page():
    """系统操作日志页面"""
    try:
        if session.get('role') != 'admin':
            return redirect(url_for('login_page'))
        
        return render_template('system-logs.html',
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/api/system-logs', methods=['GET'])
@api_login_required
def get_system_logs():
    """获取系统操作日志"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可以查看日志'}), 403
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # 分页查询
        query = SystemLog.query.order_by(SystemLog.created_at.desc())
        total = query.count()
        logs = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'status': 'success',
            'data': [log.to_dict() for log in logs.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': logs.pages
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/system-logs/filter', methods=['POST'])
@api_login_required
def filter_system_logs():
    """筛选系统操作日志"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可以查看日志'}), 403
        
        data = request.get_json()
        page = data.get('page', 1)
        per_page = data.get('per_page', 50)
        
        # 构建查询
        query = SystemLog.query
        
        # 按操作类型筛选
        if data.get('action'):
            query = query.filter_by(action=data['action'])
        
        # 按资源类型筛选
        if data.get('resource_type'):
            query = query.filter_by(resource_type=data['resource_type'])
        
        # 按用户筛选
        if data.get('username'):
            query = query.filter(SystemLog.username.ilike(f"%{data['username']}%"))
        
        # 按时间范围筛选
        if data.get('start_date'):
            try:
                start_dt = datetime.fromisoformat(data['start_date'])
                query = query.filter(SystemLog.created_at >= start_dt)
            except:
                pass
        
        if data.get('end_date'):
            try:
                end_dt = datetime.fromisoformat(data['end_date'])
                query = query.filter(SystemLog.created_at <= end_dt)
            except:
                pass
        
        # 排序和分页
        query = query.order_by(SystemLog.created_at.desc())
        total = query.count()
        logs = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'status': 'success',
            'data': [log.to_dict() for log in logs.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': logs.pages
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 设备控制日志详情查看 API (Task5)
# ============================================================

@app.route('/device-logs/<int:device_id>', methods=['GET'])
@login_required
def device_logs_page(device_id):
    """设备操作日志页面"""
    try:
        device = Device.query.get(device_id)
        if not device:
            return render_template('error.html', message='设备不存在'), 404
        
        return render_template('device-logs.html',
                               device_id=device_id,
                               device_name=device.device_name,
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/api/devices/<int:device_id>/logs', methods=['GET'])
@api_login_required
def get_device_logs(device_id):
    """获取设备操作日志"""
    try:
        # 验证设备存在
        device = Device.query.get(device_id)
        if not device:
            return jsonify({'status': 'error', 'message': '设备不存在'}), 404
        
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # 查询该设备的日志
        query = DeviceLog.query.filter_by(device_id=device_id).order_by(DeviceLog.log_time.desc())
        total = query.count()
        logs = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'status': 'success',
            'data': [{
                'id': log.id,
                'action': log.action,
                'operator': log.operator,
                'previous_state': log.previous_state,
                'current_state': log.current_state,
                'details': log.details,
                'log_time': log.log_time.isoformat() if log.log_time else None
            } for log in logs.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': logs.pages
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/devices/<int:device_id>/logs/filter', methods=['POST'])
@api_login_required
def filter_device_logs(device_id):
    """筛选设备操作日志"""
    try:
        # 验证设备存在
        device = Device.query.get(device_id)
        if not device:
            return jsonify({'status': 'error', 'message': '设备不存在'}), 404
        
        data = request.get_json()
        page = data.get('page', 1)
        per_page = data.get('per_page', 50)
        
        # 构建查询
        query = DeviceLog.query.filter_by(device_id=device_id)
        
        # 按操作类型筛选
        if data.get('action'):
            query = query.filter_by(action=data['action'])
        
        # 按操作员筛选
        if data.get('operator'):
            query = query.filter(DeviceLog.operator.ilike(f"%{data['operator']}%"))
        
        # 按时间范围筛选
        if data.get('start_time'):
            try:
                start_dt = datetime.fromisoformat(data['start_time'])
                query = query.filter(DeviceLog.log_time >= start_dt)
            except:
                pass
        
        if data.get('end_time'):
            try:
                end_dt = datetime.fromisoformat(data['end_time'])
                query = query.filter(DeviceLog.log_time <= end_dt)
            except:
                pass
        
        # 排序和分页
        query = query.order_by(DeviceLog.log_time.desc())
        total = query.count()
        logs = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'status': 'success',
            'data': [{
                'id': log.id,
                'action': log.action,
                'operator': log.operator,
                'previous_state': log.previous_state,
                'current_state': log.current_state,
                'details': log.details,
                'log_time': log.log_time.isoformat() if log.log_time else None
            } for log in logs.items],
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': logs.pages
            }
        }), 200
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 数据导出功能 (Task6)
# ============================================================

def create_export_file(export_type):
    """创建导出文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        if export_type == 'ponds':
            ws.title = '鱼塘列表'
            ws['A1'] = '鱼塘ID'
            ws['B1'] = '鱼塘名称'
            ws['C1'] = '面积(亩)'
            ws['D1'] = '位置'
            ws['E1'] = '状态'
            ws['F1'] = '创建时间'
            
            for col in ws.iter_cols(min_row=1, max_col=6, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
            
            ponds = Pond.query.all()
            for idx, pond in enumerate(ponds, 2):
                ws[f'A{idx}'] = pond.id
                ws[f'B{idx}'] = pond.pond_name
                ws[f'C{idx}'] = pond.area
                ws[f'D{idx}'] = pond.location or '-'
                ws[f'E{idx}'] = pond.status or '-'
                ws[f'F{idx}'] = pond.created_at.strftime('%Y-%m-%d %H:%M:%S') if pond.created_at else '-'
        
        elif export_type == 'devices':
            ws.title = '设备列表'
            ws['A1'] = '设备ID'
            ws['B1'] = '设备名称'
            ws['C1'] = '设备类型'
            ws['D1'] = '鱼塘'
            ws['E1'] = '状态'
            ws['F1'] = '功率消耗'
            ws['G1'] = '最后活动'
            
            for col in ws.iter_cols(min_row=1, max_col=7, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
            
            devices = Device.query.all()
            for idx, device in enumerate(devices, 2):
                pond = Pond.query.get(device.pond_id)
                ws[f'A{idx}'] = device.id
                ws[f'B{idx}'] = device.device_name
                ws[f'C{idx}'] = device.device_type
                ws[f'D{idx}'] = pond.pond_name if pond else '-'
                ws[f'E{idx}'] = device.status or '-'
                ws[f'F{idx}'] = device.power_consumption or '-'
                ws[f'G{idx}'] = device.last_active.strftime('%Y-%m-%d %H:%M:%S') if device.last_active else '-'
        
        elif export_type == 'water_quality':
            ws.title = '水质数据'
            ws['A1'] = '鱼塘'
            ws['B1'] = '温度(°C)'
            ws['C1'] = 'pH值'
            ws['D1'] = '溶氧(mg/L)'
            ws['E1'] = '盐度(%)'
            ws['F1'] = '氨氮(mg/L)'
            ws['G1'] = '亚硝酸盐(mg/L)'
            ws['H1'] = '测量时间'
            
            for col in ws.iter_cols(min_row=1, max_col=8, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
            
            # 这里需要从水质数据表中获取，假设有类似的表
            # 为了演示，我们只导出阈值配置
            thresholds = WaterQualityThreshold.query.all()
            ws.clear()
            ws.title = '水质告警阈值'
            ws['A1'] = '参数名称'
            ws['B1'] = '最小值'
            ws['C1'] = '最大值'
            ws['D1'] = '告警级别'
            ws['E1'] = '单位'
            ws['F1'] = '状态'
            
            for col in ws.iter_cols(min_row=1, max_col=6, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
            
            for idx, threshold in enumerate(thresholds, 2):
                ws[f'A{idx}'] = threshold.parameter_name
                ws[f'B{idx}'] = threshold.min_value or '-'
                ws[f'C{idx}'] = threshold.max_value or '-'
                ws[f'D{idx}'] = threshold.warning_level
                ws[f'E{idx}'] = threshold.unit or '-'
                ws[f'F{idx}'] = '启用' if threshold.is_active else '禁用'
        
        elif export_type == 'device_logs':
            ws.title = '设备操作日志'
            ws['A1'] = '设备ID'
            ws['B1'] = '操作'
            ws['C1'] = '操作员'
            ws['D1'] = '前一状态'
            ws['E1'] = '当前状态'
            ws['F1'] = '操作时间'
            
            for col in ws.iter_cols(min_row=1, max_col=6, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
            
            logs = DeviceLog.query.order_by(DeviceLog.log_time.desc()).limit(1000).all()
            for idx, log in enumerate(logs, 2):
                ws[f'A{idx}'] = log.device_id
                ws[f'B{idx}'] = log.action
                ws[f'C{idx}'] = log.operator or '-'
                ws[f'D{idx}'] = log.previous_state or '-'
                ws[f'E{idx}'] = log.current_state or '-'
                ws[f'F{idx}'] = log.log_time.strftime('%Y-%m-%d %H:%M:%S') if log.log_time else '-'
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    except ImportError:
        return None
    except Exception as e:
        print(f"[ERROR] 导出文件创建失败: {e}")
        return None


@app.route('/api/export/<export_type>', methods=['POST'])
@api_login_required
def export_data(export_type):
    """导出数据为Excel文件"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可以导出数据'}), 403
        
        # 验证导出类型
        valid_types = ['ponds', 'devices', 'water_quality', 'device_logs']
        if export_type not in valid_types:
            return jsonify({'status': 'error', 'message': '无效的导出类型'}), 400
        
        # 创建导出文件
        file_obj = create_export_file(export_type)
        if file_obj is None:
            return jsonify({'status': 'error', 'message': 'openpyxl库未安装，无法导出'}), 500
        
        # 生成文件名
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        type_names = {
            'ponds': '鱼塘列表',
            'devices': '设备列表',
            'water_quality': '水质数据',
            'device_logs': '设备日志'
        }
        filename = f'{type_names.get(export_type, export_type)}_{timestamp}.xlsx'
        
        # 记录审计日志
        record_audit_log('export', 'Data', details=f'导出{type_names.get(export_type)}')
        
        # 返回文件
        from flask import send_file
        return send_file(
            file_obj,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 登录系统优化与密码管理 (Task7)
# ============================================================

@app.route('/user-profile', methods=['GET'])
@login_required
def user_profile():
    """用户资料页面"""
    try:
        return render_template('user-profile.html',
                               username=session.get('username'),
                               role=session.get('role'))
    except Exception as e:
        return render_template('error.html', message=str(e)), 500


@app.route('/api/user/change-password', methods=['POST'])
@api_login_required
def change_password():
    """修改用户密码"""
    try:
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        # 验证输入
        if not current_password or not new_password:
            return jsonify({'status': 'error', 'message': '密码不能为空'}), 400
        
        if len(new_password) < 6:
            return jsonify({'status': 'error', 'message': '新密码长度至少为6个字符'}), 400
        
        user = User.query.get(session['user_id'])
        if not user:
            return jsonify({'status': 'error', 'message': '用户不存在'}), 404
        
        # 验证当前密码
        if user.password_hash != current_password:
            return jsonify({'status': 'error', 'message': '当前密码错误'}), 401
        
        # 更新密码
        user.password_hash = new_password
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        # 记录审计日志
        record_audit_log('update', 'User', user.id, details='用户修改密码')
        
        return jsonify({
            'status': 'success',
            'message': '密码修改成功'
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/admin/create-user', methods=['POST'])
@api_login_required
def create_user():
    """创建新用户（仅管理员）"""
    try:
        if session.get('role') != 'admin':
            return jsonify({'status': 'error', 'message': '仅管理员可以创建用户'}), 403
        
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        role = data.get('role', 'user')
        full_name = data.get('full_name')
        email = data.get('email')
        
        # 验证输入
        if not username or not password:
            return jsonify({'status': 'error', 'message': '用户名和密码不能为空'}), 400
        
        if len(password) < 6:
            return jsonify({'status': 'error', 'message': '密码长度至少为6个字符'}), 400
        
        # 检查用户名是否已存在
        if User.query.filter_by(username=username).first():
            return jsonify({'status': 'error', 'message': '用户名已存在'}), 409
        
        # 验证角色
        valid_roles = ['admin', 'user', 'supplier']
        if role not in valid_roles:
            return jsonify({'status': 'error', 'message': f'无效的角色，必须为: {", ".join(valid_roles)}'}), 400
        
        # 创建新用户
        new_user = User(
            username=username,
            password_hash=password,
            role=role,
            full_name=full_name,
            email=email,
            is_active=True,
            created_at=datetime.utcnow()
        )
        
        db.session.add(new_user)
        db.session.commit()
        
        # 记录审计日志
        record_audit_log('create', 'User', new_user.id, 
                        details=f'管理员{session.get("username")}创建新用户{username}')
        
        return jsonify({
            'status': 'success',
            'message': f'用户{username}创建成功',
            'data': {
                'id': new_user.id,
                'username': new_user.username,
                'role': new_user.role,
                'full_name': new_user.full_name,
                'email': new_user.email
            }
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# 供应商管理 API 注册（任务6）
# ============================================================

try:
    from supplier_api import register_supplier_apis
    register_supplier_apis(app, db, Supplier, SeedlingProduct, SeedlingInventory, PurchaseOrder, OrderItem, User, Pond)
    print("[OK] 供应商管理API已注册")
except ImportError as e:
    print(f"[WARN] 供应商API导入失败: {e}")
except Exception as e:
    print(f"[WARN] 供应商API注册失败: {e}")


# ============================================================
# 主程序入口
# ============================================================

if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            print('数据库表已创建或已存在')
        except Exception as e:
            print(f'创建表失败: {e}')
    
    # 调试输出：打印所有 /api/ 路由
    print('\n=== 已注册的 API 路由 ===')
    for rule in app.url_map.iter_rules():
        if '/api' in rule.rule:
            print(f'  {rule.rule} [{rule.methods}]')
    print()
    
    print('智慧渔场管理系统启动中...')
    print('访问地址: http://127.0.0.1:5000')
    print('按 Ctrl+C 停止服务')
    
    # 启动硬件采集
    with app.app_context():
        try:
            init_hardware_collection()
        except Exception as e:
            print(f"[WARN] 硬件采集启动失败: {e}")
    
    try:
        app.run(host='127.0.0.1', port=5000, debug=True, use_reloader=False)
    except KeyboardInterrupt:
        print("\n应用正在关闭...")
        shutdown_hardware()
        print("[OK] 已正确关闭")
